import openpyxl
from openpyxl.utils import get_column_letter
import os

fname = '18 TEAM DRAWS FINAL.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb['Sheet1']

print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')

# Check all cells rows 1-5 for header content
print('\n--- HEADER ROWS (1-5) ---')
for row in ws.iter_rows(min_row=1, max_row=5):
    for cell in row:
        if cell.value is not None:
            f = cell.font
            al = cell.alignment
            fill = cell.fill
            print(f'  {cell.coordinate}: val={repr(cell.value)}, '
                  f'bold={f.bold if f else None}, '
                  f'size={f.size if f else None}, '
                  f'align_h={al.horizontal if al else None}, '
                  f'align_v={al.vertical if al else None}, '
                  f'wrap={al.wrap_text if al else None}')

# Merged cells
print('\n--- MERGED CELLS ---')
for mc in ws.merged_cells.ranges:
    print(f'  {mc}')

# Images
print('\n--- IMAGES ---')
if hasattr(ws, '_images'):
    for i, img in enumerate(ws._images):
        print(f'  Image {i}: anchor={img.anchor}, ref={getattr(img, "ref", None)}')
        # Try to get image data
        if hasattr(img, '_data'):
            data = img._data()
            print(f'    Data size: {len(data)} bytes')
            ext = 'png'
            out_name = f'extracted_image_{i}.{ext}'
            with open(out_name, 'wb') as f_out:
                f_out.write(data)
            print(f'    Saved to: {out_name}')

# Column widths
print('\n--- COLUMN WIDTHS ---')
for col_letter, col_dim in ws.column_dimensions.items():
    if col_dim.width:
        print(f'  Col {col_letter}: width={col_dim.width}')

# Row 1-3 heights
print('\n--- ROW HEIGHTS (1-5) ---')
for r in range(1, 6):
    h = ws.row_dimensions[r].height
    print(f'  Row {r}: height={h}')

# Page setup
print('\n--- PAGE SETUP ---')
ps = ws.page_setup
print(f'  orientation={ps.orientation}, paperSize={ps.paperSize}')
print(f'  fitToPage={ps.fitToPage}')

# Print area
print(f'\n--- PRINT AREA ---')
print(f'  {ws.print_area}')

# Header/footer
hf = ws.oddHeader
print(f'\n--- HEADER ---')
print(f'  left={hf.left.text if hf.left else None}')
print(f'  center={hf.center.text if hf.center else None}')
print(f'  right={hf.right.text if hf.right else None}')
