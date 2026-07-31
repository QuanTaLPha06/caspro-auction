"""
Tournament Draw Excel Generator
Creates exact replicas of the reference format for Dodgeball and Kho Kho draws.

Format from reference files:
- Row 1: EMPTY (no header)
- 'cc' cells have medium borders on all 4 sides
- 'bye' and 'best loser' text are plain (no border)
- Row heights: 16.2 for most rows
- No merged cells, no column width settings
- The 'cc' placeholder represents a team slot in the draw
"""

import openpyxl
from openpyxl.styles import Font, Border, Side

def medium_border():
    medium = Side(style='medium')
    return Border(left=medium, right=medium, top=medium, bottom=medium)

def cc(ws, row, col):
    """Place a team slot cell with medium border"""
    cell = ws.cell(row=row, column=col, value='cc')
    cell.border = medium_border()
    cell.font = Font(size=12)

def txt(ws, row, col, value):
    """Place a plain text cell (no border)"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(size=12)

def set_heights(ws, rows):
    for r in rows:
        if r >= 1:
            ws.row_dimensions[r].height = 16.2


# ================================================================
# 13 TEAM DRAW - exactly matches 13 TEAM DRAWS FINAL.xlsx
# 28 cc cells, starts row 2
# Groups: uses cols A(1), C(3), E(5), G(7), J(10)
# ================================================================
def draw_13(ws):
    # Group 1: rows 2-5
    cc(ws,2,1); cc(ws,3,3); cc(ws,4,1); cc(ws,5,5)
    # Group 2: rows 7-10
    cc(ws,7,1); cc(ws,8,3); cc(ws,9,1); cc(ws,10,7)
    # Group 3: rows 12-15
    cc(ws,12,1); cc(ws,13,3); cc(ws,14,1); cc(ws,15,5)
    # Group 4: rows 17-19
    cc(ws,17,1); cc(ws,18,3); cc(ws,19,1)
    # Semi-final connector
    cc(ws,19,10)
    # Group 5: rows 22-25
    cc(ws,22,1); cc(ws,23,3); cc(ws,24,1); cc(ws,25,5)
    # Group 6: rows 27-29
    cc(ws,27,1); cc(ws,28,3); cc(ws,29,1); cc(ws,29,7)
    # Knockout
    cc(ws,32,1); cc(ws,32,3)
    txt(ws,33,2,'(bye)')
    cc(ws,34,5)
    cc(ws,36,3)
    txt(ws,38,1,'( best loser)')
    
    set_heights(ws, list(range(2, 40)))


# ================================================================
# 14 TEAM DRAW - exactly matches 14 TEAM DRAWS FINAL.xlsx
# 29 cc cells, starts col B(2)
# ================================================================
def draw_14(ws):
    # Group 1
    cc(ws,2,2); cc(ws,3,4); cc(ws,4,2); cc(ws,5,6)
    # Group 2
    cc(ws,7,2); cc(ws,8,4); cc(ws,9,2); cc(ws,10,8)
    # Group 3
    cc(ws,12,2); cc(ws,13,4); cc(ws,14,2); cc(ws,15,6)
    # Group 4
    cc(ws,17,2); cc(ws,18,4); cc(ws,19,2)
    cc(ws,20,11)
    # Group 5
    cc(ws,22,2); cc(ws,23,4); cc(ws,24,2); cc(ws,25,6)
    # Group 6
    cc(ws,27,2); cc(ws,28,4); cc(ws,29,2); cc(ws,30,8)
    # Group 7
    cc(ws,32,2); cc(ws,33,4); cc(ws,34,2)
    # Knockout
    cc(ws,36,6)
    cc(ws,38,4)
    txt(ws,39,4,'best loser')
    
    set_heights(ws, list(range(2, 42)))


# ================================================================
# 15 TEAM DRAW - exactly matches 15 TEAM DRAWS FINAL.xlsx
# 30 cc cells, starts col A, uses col L(12) for connector
# ================================================================
def draw_15(ws):
    # Group 1
    cc(ws,2,1); cc(ws,3,3); cc(ws,4,1); cc(ws,5,5)
    # Group 2
    cc(ws,7,1); cc(ws,8,3); cc(ws,9,1); cc(ws,10,8)
    # Group 3
    cc(ws,12,1); cc(ws,13,3); cc(ws,14,1); cc(ws,15,5)
    # Group 4
    cc(ws,17,1); cc(ws,18,3); cc(ws,19,1)
    cc(ws,20,12)
    # Group 5
    cc(ws,22,1); cc(ws,23,3); cc(ws,24,1); cc(ws,25,5)
    # Group 6
    cc(ws,27,1); cc(ws,28,3); cc(ws,29,1); cc(ws,30,8)
    # Group 7
    cc(ws,32,1); cc(ws,33,3); cc(ws,34,1); cc(ws,35,5)
    # Knockout (bye match)
    cc(ws,37,1); cc(ws,37,3)
    txt(ws,38,2,'( bye)')
    
    set_heights(ws, list(range(2, 40)))


# ================================================================
# 16 TEAM DRAW - exactly matches 16 TEAM DRAWS FINAL.xlsx
# 31 cc cells, starts row 3 (rows 1-2 empty)
# ================================================================
def draw_16(ws):
    # Group 1
    cc(ws,3,1); cc(ws,4,3); cc(ws,5,1); cc(ws,6,5)
    # Group 2
    cc(ws,8,1); cc(ws,9,3); cc(ws,10,1); cc(ws,11,8)
    # Group 3
    cc(ws,13,1); cc(ws,14,3); cc(ws,15,1); cc(ws,16,5)
    # Group 4
    cc(ws,18,1); cc(ws,19,3); cc(ws,20,1); cc(ws,21,11)
    # Group 5
    cc(ws,23,1); cc(ws,24,3); cc(ws,25,1); cc(ws,26,5)
    # Group 6
    cc(ws,28,1); cc(ws,29,3); cc(ws,30,1); cc(ws,31,8)
    # Group 7
    cc(ws,33,1); cc(ws,34,3); cc(ws,35,1); cc(ws,36,5)
    # Group 8
    cc(ws,38,1); cc(ws,39,3); cc(ws,40,1)
    
    set_heights(ws, list(range(3, 43)))


# ================================================================
# 17 TEAM DRAW - exactly matches 17 TEAM DRAWS FINAL.xlsx
# 40 cc cells, starts row 3
# ================================================================
def draw_17(ws):
    # Group 1 (3 teams)
    cc(ws,3,1); cc(ws,4,3); cc(ws,5,1)
    cc(ws,7,5)
    # Group 2 (3 teams)
    cc(ws,8,1); cc(ws,9,3); cc(ws,10,1)
    cc(ws,12,7)
    # Group 3 (3 teams)
    cc(ws,13,1); cc(ws,14,3); cc(ws,15,1)
    cc(ws,17,5)
    # Group 4 (3 teams)
    cc(ws,18,1); cc(ws,19,3); cc(ws,20,1)
    cc(ws,21,9)
    # Group 5 (4 teams)
    cc(ws,23,1); cc(ws,24,3); cc(ws,25,1); cc(ws,26,5)
    # Group 6 (4 teams)
    cc(ws,28,1); cc(ws,29,3); cc(ws,30,1); cc(ws,31,7)
    # Group 7 (4 teams)
    cc(ws,33,1); cc(ws,34,3); cc(ws,35,1); cc(ws,36,5)
    # Group 8 connector
    cc(ws,38,1); cc(ws,38,12)
    cc(ws,39,3); cc(ws,40,1)
    # Knockout stage
    cc(ws,43,1); cc(ws,43,3)
    txt(ws,45,2,'( bye)')
    cc(ws,46,5)
    cc(ws,48,3)
    txt(ws,50,3,'( best loser)'); cc(ws,50,7)
    cc(ws,54,5)
    txt(ws,56,5,'best loser'); cc(ws,56,9)
    cc(ws,62,7)
    txt(ws,64,7,'best loser')
    
    set_heights(ws, list(range(3, 68)))


# ================================================================
# 18 TEAM DRAW - exactly matches 18 TEAM DRAWS FINAL.xlsx
# 41 cc cells, starts row 3
# ================================================================
def draw_18(ws):
    # Group 1
    cc(ws,3,1); cc(ws,4,3); cc(ws,5,1); cc(ws,6,5)
    # Group 2
    cc(ws,8,1); cc(ws,9,3); cc(ws,10,1); cc(ws,11,7)
    # Group 3
    cc(ws,13,1); cc(ws,14,3); cc(ws,15,1); cc(ws,16,5)
    # Group 4
    cc(ws,18,1); cc(ws,19,3); cc(ws,20,1); cc(ws,21,9)
    # Group 5
    cc(ws,23,1); cc(ws,24,3); cc(ws,25,1); cc(ws,26,5)
    # Group 6
    cc(ws,28,1); cc(ws,29,3); cc(ws,30,1); cc(ws,31,7)
    # Group 7
    cc(ws,33,1); cc(ws,34,3); cc(ws,35,1); cc(ws,36,5)
    # Group 8
    cc(ws,38,1); cc(ws,39,3); cc(ws,40,1); cc(ws,40,11)
    # Group 9
    cc(ws,43,1); cc(ws,44,3); cc(ws,45,1); cc(ws,46,5)
    # Knockout
    cc(ws,49,3)
    txt(ws,51,3,'best loser'); cc(ws,51,7)
    cc(ws,56,5)
    txt(ws,58,5,'best loser')
    cc(ws,59,9)
    cc(ws,67,7)
    txt(ws,69,7,'best loser')
    
    set_heights(ws, list(range(3, 73)))


# ================================================================
# 19 TEAM DRAW - exactly matches 19 TEAM DRAWS FINAL.xlsx
# 42 cc cells, starts row 3
# ================================================================
def draw_19(ws):
    # Group 1
    cc(ws,3,1); cc(ws,4,3); cc(ws,5,1); cc(ws,6,5)
    # Group 2
    cc(ws,8,1); cc(ws,9,3); cc(ws,10,1); cc(ws,11,7)
    # Group 3
    cc(ws,13,1); cc(ws,14,3); cc(ws,15,1); cc(ws,16,5)
    # Group 4
    cc(ws,18,1); cc(ws,19,3); cc(ws,20,1); cc(ws,21,9)
    # Group 5 (3 teams)
    cc(ws,23,1); cc(ws,24,3); cc(ws,25,1)
    cc(ws,27,5)
    # Group 6 (3 teams)
    cc(ws,28,1); cc(ws,29,3); cc(ws,30,1)
    cc(ws,32,7)
    # Group 7 (3 teams)
    cc(ws,33,1); cc(ws,34,3); cc(ws,35,1)
    cc(ws,37,5)
    # Group 8 connector
    cc(ws,38,1); cc(ws,38,12)
    cc(ws,39,3); cc(ws,40,1)
    # Group 9
    cc(ws,43,1); cc(ws,44,3); cc(ws,45,1); cc(ws,46,5)
    # Knockout with bye
    cc(ws,48,1); cc(ws,48,3)
    txt(ws,49,2,'bye'); cc(ws,49,7)
    cc(ws,53,5)
    txt(ws,55,5,'best loser'); cc(ws,55,9)
    cc(ws,61,7)
    txt(ws,63,7,'best loser')
    
    set_heights(ws, list(range(3, 67)))


# ================================================================
# 11 TEAM DRAW (new - derived from pattern)
# 3 groups of 3 + 1 group of 2 = 11 teams
# ================================================================
def draw_11(ws):
    # Group 1 (3 teams)
    cc(ws,2,1); cc(ws,3,3); cc(ws,4,1); cc(ws,5,5)
    # Group 2 (3 teams)
    cc(ws,7,1); cc(ws,8,3); cc(ws,9,1); cc(ws,10,7)
    # Group 3 (3 teams)
    cc(ws,12,1); cc(ws,13,3); cc(ws,14,1); cc(ws,15,5)
    # Group 4 (2 teams)
    cc(ws,17,1); cc(ws,18,3); cc(ws,19,1)
    cc(ws,19,9)
    # Knockout
    cc(ws,22,1); cc(ws,22,3)
    txt(ws,23,2,'( bye)')
    cc(ws,24,5)
    cc(ws,26,3)
    txt(ws,28,1,'( best loser)')
    
    set_heights(ws, list(range(2, 31)))


# ================================================================
# 12 TEAM DRAW (new - derived from pattern)
# 4 groups of 3 teams each
# ================================================================
def draw_12(ws):
    # Group 1 (3 teams)
    cc(ws,2,1); cc(ws,3,3); cc(ws,4,1); cc(ws,5,5)
    # Group 2 (3 teams)
    cc(ws,7,1); cc(ws,8,3); cc(ws,9,1); cc(ws,10,7)
    # Group 3 (3 teams)
    cc(ws,12,1); cc(ws,13,3); cc(ws,14,1); cc(ws,15,5)
    # Group 4 (3 teams)
    cc(ws,17,1); cc(ws,18,3); cc(ws,19,1); cc(ws,20,9)
    # Semi connector
    cc(ws,22,1); cc(ws,23,3); cc(ws,24,1)
    cc(ws,26,5)
    # Final
    cc(ws,28,3)
    
    set_heights(ws, list(range(2, 31)))


# ================================================================
# GENERATE ALL FILES
# ================================================================
import os

def make_draw(filename, draw_fn, sport_name, num_teams):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    # Row 1: Sport name header (matching the user's request to have sport name written)
    header = ws.cell(row=1, column=1, value=f'{sport_name}')
    header.font = Font(size=14, bold=True)
    draw_fn(ws)
    wb.save(filename)
    print(f'Created: {filename}')

# DODGEBALL draws
make_draw('DODGEBALL 11 TEAM DRAWS FINAL.xlsx', draw_11, 'DODGEBALL', 11)
make_draw('DODGEBALL 12 TEAM DRAWS FINAL.xlsx', draw_12, 'DODGEBALL', 12)
make_draw('DODGEBALL 13 TEAM DRAWS FINAL.xlsx', draw_13, 'DODGEBALL', 13)
make_draw('DODGEBALL 14 TEAM DRAWS FINAL.xlsx', draw_14, 'DODGEBALL', 14)
make_draw('DODGEBALL 15 TEAM DRAWS FINAL.xlsx', draw_15, 'DODGEBALL', 15)
make_draw('DODGEBALL 16 TEAM DRAWS FINAL.xlsx', draw_16, 'DODGEBALL', 16)

# KHO KHO draws
make_draw('KHO KHO 14 TEAM DRAWS FINAL.xlsx', draw_14, 'KHO KHO', 14)
make_draw('KHO KHO 15 TEAM DRAWS FINAL.xlsx', draw_15, 'KHO KHO', 15)
make_draw('KHO KHO 16 TEAM DRAWS FINAL.xlsx', draw_16, 'KHO KHO', 16)
make_draw('KHO KHO 17 TEAM DRAWS FINAL.xlsx', draw_17, 'KHO KHO', 17)
make_draw('KHO KHO 18 TEAM DRAWS FINAL.xlsx', draw_18, 'KHO KHO', 18)
make_draw('KHO KHO 19 TEAM DRAWS FINAL.xlsx', draw_19, 'KHO KHO', 19)

print('\nAll 12 draw files generated successfully!')
