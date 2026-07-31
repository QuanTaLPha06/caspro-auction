import openpyxl
from openpyxl.utils import get_column_letter

# Deep inspect ALL cells including those with borders but no value
# to find the bracket connector lines

def inspect_all_borders(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb['Sheet1']
    
    print(f'\n=== {fname} ===')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    print('All cells with ANY border:')
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            b = cell.border
            border_parts = []
            if b:
                if b.top and b.top.style:
                    border_parts.append(f'T:{b.top.style}')
                if b.bottom and b.bottom.style:
                    border_parts.append(f'B:{b.bottom.style}')
                if b.left and b.left.style:
                    border_parts.append(f'L:{b.left.style}')
                if b.right and b.right.style:
                    border_parts.append(f'R:{b.right.style}')
            
            if border_parts or cell.value is not None:
                val = repr(cell.value) if cell.value is not None else '(empty)'
                print(f'  {cell.coordinate}: {val} [{", ".join(border_parts)}]')

for fname in ['13 TEAM DRAWS FINAL.xlsx']:
    inspect_all_borders(fname)
