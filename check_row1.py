import openpyxl
from openpyxl.utils import get_column_letter

# Check ALL cells including row 1 for all files
files = [
    '13 TEAM DRAWS FINAL.xlsx',
    '14 TEAM DRAWS FINAL.xlsx', 
    '15 TEAM DRAWS FINAL.xlsx',
    '16 TEAM DRAWS FINAL.xlsx',
    '17 TEAM DRAWS FINAL.xlsx',
    '18 TEAM DRAWS FINAL.xlsx',
    '19 TEAM DRAWS FINAL.xlsx',
]

for fname in files:
    print(f'\n=== {fname} ===')
    wb = openpyxl.load_workbook(fname)
    ws = wb['Sheet1']
    
    # Print row 1 details
    print('Row 1 content:')
    for cell in ws[1]:
        if cell.value is not None:
            print(f'  {cell.coordinate}: {repr(cell.value)}')
    
    # Print first data row
    print(f'First data row with cc: row 2 or 3')
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value:
                print(f'  {cell.coordinate}: {repr(cell.value)}')
    print()
    
    # Count total cc cells
    cc_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'cc' or cell.value == 'CC':
                cc_count += 1
    print(f'Total cc cells: {cc_count}')
